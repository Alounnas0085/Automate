"""
cegid_to_monplanning.py
-----------------------
Convertit un fichier xlsx extrait de Cegid en CSV Mon-Planning.

- Copie l'original dans Conversions/original_TIMESTAMP.xlsx
- Génère Conversions/monplanning_TIMESTAMP.csv

Colonnes Cegid   → Colonnes Mon-Planning
-----------------------------------------
Date             → debut + fin  (groupement jours consécutifs)
Client           → client
Produit          → type  (via TYPE_MAP)
Mission          → intitule
HeureDebut/Fin   → planification  (journee complète / matin / apres-midi)
config.consultant→ consultant
"""

import csv
import json
import re
import shutil
from datetime import date, timedelta
from pathlib import Path

import openpyxl

# ── Chemins ────────────────────────────────────────────────
BASE_DIR       = Path(__file__).resolve().parent.parent
EXTRACTIONS    = BASE_DIR / "Extractions"
CONVERSIONS    = BASE_DIR / "Conversions"
CONFIG_PATH    = Path(__file__).resolve().parent / "config.json"

CONVERSIONS.mkdir(parents=True, exist_ok=True)

# ── Mapping type Cegid → type Mon-Planning ─────────────────
TYPE_MAP = {
    "tma"        : "TMA",
    "bpo"        : "BPO",
    "projet"     : "Projet",
    "project"    : "Projet",
    "migration"  : "Projet",
    "migrations" : "Projet",
    "formation"  : "Formation",
    "regie"      : "Régie",
    "régie"      : "Régie",
}

# ── Mois français ──────────────────────────────────────────
MOIS = {
    "janvier":1,"février":2,"mars":3,"avril":4,"mai":5,"juin":6,
    "juillet":7,"août":8,"septembre":9,"octobre":10,"novembre":11,"décembre":12,
    "fevrier":2,"aout":8,  # variantes sans accent
}


def load_config():
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def timestamp():
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def parse_date(value) -> date | None:
    """Accepte un objet datetime/date openpyxl ou un texte dans plusieurs formats."""
    from datetime import datetime as dt
    if isinstance(value, dt):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text or text.lower() in ("none", ""):
        return None
    # DD-MM-YYYY ou DD/MM/YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", text)
    if m:
        day, month, year = m.groups()
        try:
            return date(int(year), int(month), int(day))
        except ValueError:
            pass
    # YYYY-MM-DD (openpyxl datetime converti en str)
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        year, month, day = m.groups()
        try:
            return date(int(year), int(month), int(day))
        except ValueError:
            pass
    # Texte français : 'Lundi 21 mars 2026'
    text_low = text.lower()
    m = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})", text_low)
    if m:
        day, month_str, year = m.groups()
        month = MOIS.get(month_str)
        if month:
            try:
                return date(int(year), month, int(day))
            except ValueError:
                pass
    return None


def map_type(produit: str, type_projet: str) -> str:
    """Déduit le type Mon-Planning depuis Produit / TypeProjet."""
    for src in (produit, type_projet):
        key = (src or "").strip().lower()
        if key in TYPE_MAP:
            return TYPE_MAP[key]
        for k, v in TYPE_MAP.items():
            if k in key:
                return v
    return (produit or "TMA").strip() or "TMA"


def get_planification(h_debut: str, h_fin: str) -> str:
    """
    Déduit la planification depuis les heures Cegid.
    matin            : début < 12 h ET fin ≤ 13 h
    apres-midi       : début ≥ 12 h
    journee complète : début < 12 h ET fin > 13 h
    """
    if not h_debut:
        return "journee complète"
    try:
        hd = int(h_debut.split(":")[0])
        hf = int(h_fin.split(":")[0]) if h_fin else 24
        if hd < 12 and hf <= 13:
            return "matin"
        if hd >= 12:
            return "apres-midi"
        return "journee complète"
    except (ValueError, IndexError):
        pass
    return "journee complète"


# Noms alternatifs acceptés pour chaque colonne attendue
COLUMN_ALIASES = {
    "Date"       : ["Date", "date", "DATE", "Jour", "jour"],
    "Produit"    : ["Produit", "produit", "PRODUIT", "Product"],
    "Mission"    : ["Mission", "mission", "MISSION", "Intitulé", "Intitule"],
    "TypeProjet" : ["TypeProjet", "Type Projet", "Type", "type", "TypeMission"],
    "Client"     : ["Client", "client", "CLIENT", "Société", "Societe"],
    "HeureDebut" : ["HeureDebut", "Heure Debut", "Debut", "HeureD", "Heure_Debut"],
    "HeureFin"   : ["HeureFin", "Heure Fin", "Fin", "HeureF", "Heure_Fin"],
}


def resolve_headers(headers: list[str]) -> dict[str, int]:
    """Résout les colonnes même si les noms varient d'un fichier à l'autre."""
    raw_idx = {h: i for i, h in enumerate(headers)}
    resolved = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in raw_idx:
                resolved[canonical] = raw_idx[alias]
                break
    return resolved


def read_xlsx(path: Path) -> list[dict]:
    """Lit le xlsx Cegid et retourne une liste de dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(c).strip() if c else "" for c in rows[0]]
    idx = resolve_headers(headers)

    needed = {"Date", "Produit", "Mission", "TypeProjet", "Client", "HeureDebut", "HeureFin"}
    missing = needed - set(idx.keys())
    if missing:
        print(f"  Colonnes trouvées : {[h for h in headers if h]}")
        raise ValueError(f"Colonnes manquantes : {missing}")

    data = []
    for row in rows[1:]:
        d = parse_date(row[idx["Date"]])
        if not d:
            continue
        data.append({
            "date"       : d,
            "produit"    : str(row[idx["Produit"]]    or "").strip(),
            "mission"    : str(row[idx["Mission"]]    or "").strip(),
            "type_projet": str(row[idx["TypeProjet"]] or "").strip(),
            "client"     : str(row[idx["Client"]]     or "").strip(),
            "h_debut"    : str(row[idx["HeureDebut"]] or "").strip(),
            "h_fin"      : str(row[idx["HeureFin"]]   or "").strip(),
        })

    wb.close()
    return data


CONVERTED_COLS = {"ref", "client", "consultant", "type", "intitule", "debut", "fin"}


def read_xlsx_converted(path: Path, consultant: str) -> list[dict]:
    """Lit un xlsx déjà au format modele_missions et retourne des missions directement."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    missions = []
    for row in rows[1:]:
        debut = str(row[idx["debut"]] or "").strip()
        if not debut:
            continue
        missions.append({
            "client"       : str(row[idx["client"]]     or "").strip(),
            "consultant"   : str(row[idx.get("consultant", -1)] or consultant).strip() or consultant,
            "type"         : str(row[idx["type"]]       or "").strip(),
            "intitule"     : str(row[idx["intitule"]]   or "").strip(),
            "debut"        : debut,
            "fin"          : str(row[idx["fin"]]        or debut).strip(),
            "planification": "journee complète",
        })
    return missions


def group_missions(rows: list[dict], consultant: str) -> list[dict]:
    """
    Regroupe les lignes Cegid (1 par jour) en missions (debut → fin).
    Deux lignes appartiennent à la même mission si :
      - même (client, intitule, type, planification)
      - dates consécutives (écart ≤ 3 jours pour absorber les week-ends)
    """
    # Tri chronologique par (client, mission, date)
    rows = sorted(rows, key=lambda r: (r["client"], r["mission"], r["date"]))

    missions = []
    for r in rows:
        typ    = map_type(r["produit"], r["type_projet"])
        planif = get_planification(r["h_debut"], r["h_fin"])
        key    = (r["client"], r["mission"], typ, planif)

        if missions:
            last = missions[-1]
            gap  = (r["date"] - last["_fin"]).days
            if last["_key"] == key and gap <= 3:
                last["_fin"] = r["date"]
                continue

        missions.append({
            "_key"        : key,
            "_fin"        : r["date"],
            "client"      : r["client"],
            "consultant"  : consultant,
            "type"        : typ,
            "intitule"    : r["mission"],
            "debut"       : r["date"].strftime("%d/%m/%Y"),
            "fin"         : r["date"].strftime("%d/%m/%Y"),
            "planification": planif,
        })

    # Nettoyer les clés internes
    for m in missions:
        del m["_key"]
        fin_date = m.pop("_fin")
        m["fin"] = fin_date.strftime("%d/%m/%Y")

    return missions


def write_csv(missions: list[dict], path: Path):
    """Écrit le CSV Mon-Planning (séparateur ;, encodage UTF-8 BOM)."""
    COLS = ["ref", "client", "consultant", "type", "intitule", "debut", "fin", "planification"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLS, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        for i, m in enumerate(missions, start=1):
            writer.writerow({**{"ref": f"M{i:03d}"}, **m})
    print(f"CSV généré : {path}")


def pick_xlsx() -> list[Path]:
    """Propose les xlsx disponibles et retourne la liste des fichiers choisis."""
    files = sorted(EXTRACTIONS.glob("planning_import_*.xlsx"), reverse=True)
    if not files:
        raise FileNotFoundError(f"Aucun fichier xlsx dans {EXTRACTIONS}")

    print("\nFichiers disponibles :")
    for i, f in enumerate(files):
        print(f"  [{i+1}] {f.name}")

    print("\n  [0] Tous les fichiers")
    choice = input("\nChoisir (ex: 1  ou  1,3  ou  0 pour tout) [1] : ").strip()

    if not choice or choice == "1":
        return [files[0]]
    if choice == "0":
        return list(files)

    selected = []
    for part in choice.split(","):
        part = part.strip()
        if part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < len(files):
                selected.append(files[idx])
    return selected or [files[0]]


def deduplicate(missions: list[dict]) -> list[dict]:
    """Supprime les doublons sur (client, type, intitule, debut)."""
    seen = set()
    unique = []
    for m in missions:
        key = (m["client"], m["type"], m["intitule"], m["debut"])
        if key not in seen:
            seen.add(key)
            unique.append(m)
    return unique


def main():
    cfg        = load_config()
    consultant = cfg.get("consultant", "")
    if not consultant:
        consultant = input("Nom du consultant : ").strip()

    xlsx_paths = pick_xlsx()
    ts         = timestamp()

    # ── Lecture et fusion de toutes les extractions ────────
    all_rows      = []
    pre_missions  = []
    ok, skipped   = [], []

    for xlsx_path in xlsx_paths:
        print(f"Lecture de {xlsx_path.name}...")
        try:
            # Détecter le format du fichier
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            first_row = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True), ())
            wb.close()
            headers = {str(c).strip() for c in first_row if c}

            if CONVERTED_COLS.issubset(headers):
                # Fichier déjà converti → lire directement comme missions
                pre_missions.extend(read_xlsx_converted(xlsx_path, consultant))
            else:
                # Fichier brut Cegid → lire comme lignes journalières
                all_rows.extend(read_xlsx(xlsx_path))
            ok.append(xlsx_path.name)
        except Exception as e:
            print(f"  ⚠️  Ignoré ({e})")
            skipped.append(xlsx_path.name)

    if not all_rows and not pre_missions:
        print("\n❌ Aucune donnée valide trouvée.")
        return

    missions = group_missions(all_rows, consultant) + pre_missions
    missions = deduplicate(missions)

    csv_path = CONVERSIONS / f"monplanning_{ts}.csv"
    write_csv(missions, csv_path)

    print(f"\n✅ {len(missions)} mission(s) exportée(s) depuis {len(ok)} fichier(s)")
    print(f"   Planning  → {csv_path.name}")
    if skipped:
        print(f"\n⚠️  Fichiers ignorés ({len(skipped)}) :")
        for s in skipped:
            print(f"   - {s}")


if __name__ == "__main__":
    main()
